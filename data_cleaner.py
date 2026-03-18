import pandas as pd
from openpyxl import load_workbook

def excel_data_clean(input_file, output_file):
    """
    Метод, удаляющий данные с листа после 6 строки
    """
    print(f" Загрузка файла: {input_file}")
    workbook = load_workbook(input_file, data_only=True)

    print(f" Найдено листов: {len(workbook.sheetnames)}")

    # Создание нового Excel файл для записи
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Копирование данных из первых 6 строк
            data = []
            for row_idx in range(1, 7):  # строки 1-6
                row_data = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                data.append(row_data)

            # Создание DataFrame
            df = pd.DataFrame(data)

            # Удаление полностью пустых строки
            df = df.dropna(how='all')

            # Сохранение в новый файл
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    print(f" Результат сохранен в: {output_file}")

if __name__ == "__main__":
    excel_data_clean("Start.xlsx", "Cleaned Data.xlsx")